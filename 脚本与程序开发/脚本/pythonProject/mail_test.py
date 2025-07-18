"""
mail_test.py

author: shen
date : 2023/8/17
comment :
"""
# !/usr/local/bin/python3
import json
import os, sys
import re
import openpyxl
import time
import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.header import Header
import smtplib
import pymysql
import requests


class send_mail:
    def __init__(self, db_type, source_db):
        self.date = time.strftime("%Y%m%d", time.localtime())
        self.week = time.strftime("%w", time.localtime())
        self.day = time.strftime("%d", time.localtime())
        self.yday = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y%m%d')
        self.start_time = datetime.datetime.now()
        self.db_type = db_type
        if db_type == 'mysql':
            source_db['charset'] = 'utf8mb4'
            self.dbh = pymysql.connect(**source_db)
            self.sth = self.dbh.cursor()

    def sql_state(self, sql, mail_name, _id=0):
        try:
            info = self.sth.execute(sql)
            data_info = self.sth.fetchall()
            return data_info
        except Exception as e:
            # 输出sql执行失败告警和日志
            wechat_alert(12, mail_title=mail_name)
            sub_mail_name = mail_name + '-' + str(_id)
            log_into_db(sub_mail_name, f'sql执行失败 sql:{sql} error:{str(e)}', db_conf=mail_config, event_type=12)
            mail_type_change(_id, event_type=12, db_conf=mail_config)
            self.sth.execute('rollback')
            return -1

    def excel_info(self, excel_name, sheet_name, message_head, message_info):
        excel_path = "/tmp/mail_sql_backup/%s.xlsx" % (excel_name)
        # excel_path = os.getcwd() +"\%s.xlsx" % (excel_name)
        if os.path.isfile(excel_path):
            wb = openpyxl.load_workbook(excel_path)
        else:
            wb = openpyxl.Workbook()
            try:
                del wb["Sheet"]
            except:
                pass

        try:
            del wb['%s' % (sheet_name.strip())]
        except:
            pass
        wbs = wb.create_sheet(sheet_name)

        message_info.insert(0, message_head)
        col_max_num = len(message_head)
        row_max_num = len(message_info)

        alig = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=False, shrink_to_fit=False)

        for row_num in range(1, row_max_num + 1):
            wbs.row_dimensions[row_num].height = 20
            for col_num in range(1, col_max_num + 1):
                if row_num == 1:
                    col = openpyxl.utils.get_column_letter(col_num)
                wbs.column_dimensions[col].width = 25
                wbs.cell(row=row_num, column=col_num).value = message_info[row_num - 1][col_num - 1]
                wbs.cell(row=row_num, column=col_num).style = 'Normal'
                wbs.cell(row=row_num, column=col_num).alignment = alig

        wb.save(filename=excel_path)

        return excel_path

    def send_mail(self, mail_info):
        for mail_name, mail_sql, mail_user, mail_cc, sub_title, _id, is_backup, backup_table_name in mail_info:
            # 设定邮件时间格式
            flag = 1
            excel_flag = 0
            msg = '数据查询完成，请查收'
            # execl 列表
            excel_list = []

            mail_name = u'%s' % (re.sub(r'$', str(self.date), str(mail_name)))
            if mail_sql:
                sub = re.split(r'%s' % (str(',')), str(sub_title).strip())
                sql_text = re.split(r'%s' % (str(';')), str(mail_sql).strip())
                for sql_id in range(0, len(sql_text) - int(1)):
                    sub_mail_name = mail_name + '-' + str(sql_id)
                    info = self.sql_state(sql_text[sql_id], mail_name, sql_id)

                    # mail_sql结果数据库备份
                    if is_backup == 1:
                        mail_sql_backup(sql_text[sql_id], db_conf=mail_config, backup_table_name=backup_table_name,
                                        mail_title=sub_mail_name, mail_description=self.sth.description,
                                        data_info=list(info), id=str(_id) + '_' + str(sql_id))
                    if info == -1:
                        flag = 2
                        break

                    if flag == 1:
                        # 输出sql执行成功日志
                        log_into_db(sub_mail_name, f'sql执行成功 sql:{sql_text[sql_id]}', db_conf=mail_config,
                                    event_type=2)
                        mail_type_change(_id, mail_title=mail_name, db_conf=mail_config, event_type=2)

                        message_head = []
                        for line in self.sth.description:
                            message_head.append(line[0])
                        try:
                            attachment_name = sub[sql_id] + '_' + str(self.yday) if sub is not None else mail_name
                            excel_path = self.excel_info(attachment_name, attachment_name, list(message_head),
                                                         list(info))
                            excel_list.append(excel_path)
                            excel_flag = 1
                            log_into_db(sub_mail_name, f'execl生成成功 文件路径:{excel_path}, ',
                                        db_conf=mail_config,
                                        event_type=3)
                            mail_type_change(_id, mail_title=mail_name, db_conf=mail_config, event_type=3)
                        except Exception as e:
                            log_into_db(sub_mail_name, f'execl生成失败 error:{str(e)}',
                                        db_conf=mail_config,
                                        event_type=13)
                            mail_type_change(_id, mail_title=mail_name, db_conf=mail_config, event_type=13)
                            wechat_alert(13)

                if flag == 1:
                    msg = """<html>
                             <head>
                                <style type=\"text/css\"> body {font:13pt Arial,Helvetica,sans-serif; color:black; background:White;}
                                p{font:12pt Arial,Helvetica,sans-serif; color:black; background:White;}
                                table,tr,td {font:11pt Arial,Helvetica,sans-serif;white-space: nowrap; color:Black; background:#FCF8F8;}
                                th {font:bold 10pt Arial,Helvetica,sans-serif; color:black; background:#D8D8D8;}
                                </style>
                             </head>
                              <body>
                                    %s
                              </body>
                             </html>
                          """ % (msg)

                    mime = MIMEMultipart()
                    mime.attach(MIMEText(msg, 'html', 'utf-8'))

                    print(excel_list)
                    if excel_flag > 0:
                        # 构造附件
                        for i, sub_attach in enumerate(sub):
                            att = MIMEApplication(open(excel_list[i], 'rb').read())
                            att.add_header('Content-Disposition', 'attachment',
                                           filename=Header('%s.xlsx' % (sub[i]), 'utf-8').encode())
                            mime.attach(att)

                    mime['To'] = mail_user
                    mime['Cc'] = mail_cc
                    mime['Subject'] = mail_name
                    mime['From'] = 'DBA@dazhuanjia.com'
                    mail_user = mail_user.split(',')
                    try:
                        if mail_cc:
                            mail_user.extend(mail_cc.split(','))
                        send = smtplib.SMTP("localhost")
                        send.sendmail('localhost', mail_user, mime.as_string())
                        log_into_db(mail_name, f'邮件发送成功 附件:{sub} mail_to:{mail_user}', db_conf=mail_config,
                                    event_type=4)
                        mail_type_change(_id, mail_title=mail_name, db_conf=mail_config, event_type=4, event_end=1)
                    except Exception as e:
                        log_into_db(mail_name, f'邮件发送失败 error:{str(e)}', db_conf=mail_config, event_type=14)
                        mail_type_change(_id, mail_title=mail_name, db_conf=mail_config, event_type=14)
                        wechat_alert(14, mail_title=mail_name)
                        break


def mail_type_change(sql_id, *, mail_title=None, db_conf=None, event_type=None, event_end=None):
    """修改mail_sql 表状态

    :param : event_end: 事件是否结束 1:是 0:否
    """

    try:
        with pymysql.connect(**db_conf) as conn:
            sql = f'update mail_sql_conf set event_status = {event_type} where id = {sql_id}'
            if event_end:
                sql = f'update mail_sql_conf set event_status = {event_type} , sql_last_exec_date = "{time.strftime("%Y-%m-%d")}" where id = {sql_id}'
            print(sql)
            cursor = conn.cursor()
            cursor.execute(sql)
            conn.commit()
    except pymysql.err.Error as e:
        wechat_alert(20, extra_content=f'修改状态错误: ' + str(e), mail_title=mail_title)


def log_into_db(mail_title, context, *, db_conf=None, event_type=None):
    """ 日志输出到 MAil_log 表

    :param mail_title: 邮件标题
    :param context: 日志内容
    :param db_conf: 数据库配置
    :param event_type: 事件类型，1:获取到了sql,开始邮件事件，开始执行sql,2:sql执行成功,生成execl,3:生成execl成功，发送邮件,4:邮件发送成功，事件结束,11:获取sql列表失败,12:sql执行失败,13:execl生成失败,14:邮件发送失败,50:本次邮件事件结束
    """
    date_time = datetime.datetime.now()
    context = str(context).replace('"', '""')
    try:
        with pymysql.connect(**db_conf) as conn:
            sql = f'insert mail_log(mail_title,log_time,event_type,event_log) values ("{mail_title}",date_format("{date_time}","%Y-%m-%d %H:%i:%s"),{event_type},"{context}")'
            print(sql)
            cursor = conn.cursor()
            cursor.execute(sql)
            conn.commit()
    except pymysql.err.Error as e:
        wechat_alert(20, extra_content=f'写日志出现错误 error: ' + str(e), mail_title=mail_title)


def wechat_alert(alert_type, *, mail_title=None, extra_content=''):
    """
    :param alert_type: 4:邮件发送成功，事件结束,11:获取sql列表失败,12:sql执行失败,13:execl生成失败,14:邮件发送失败
    :param mail_title:  邮件标题
    :param extra_content: 附加内容
    :return:
    """

    content = ''
    Subject = '邮件事件失败\n' if int(alert_type) > 10 else '邮件事件成功\n'
    if mail_title:
        content += f'邮件标题: {mail_title}\n'
    content += '告警内容:'
    if alert_type == 4:
        content = content + f'本次邮件发送事件结束 \n'
    elif alert_type == 11:
        content += f'邮件发送事件失败,获取mail_sql失败 \n'
    elif alert_type == 12:
        content += f'邮件发送事件失败，查询sql执行 \n'
    elif alert_type == 13:
        content += f'邮件发送事件失败，生成execl失败 \n'
    elif alert_type == 14:
        content += f'邮件发送事件失败，发送邮件失败 \n'
    elif alert_type == 20:
        content += f'邮件发送事件失败，记录数据库日志失败\n'
    elif alert_type == 21:
        content += f'邮件发送事件失败，数据库备份失败 \n'
    elif alert_type == 22:
        content += f'邮件发送事件失败，修改sql状态失败 \n'
    elif alert_type == 23:
        content += f'邮件发送事件失败，初始化邮件状态失败 \n'
    elif alert_type == 24:
        content += f'邮件发送事件失败，备份mail_sql结果失败 \n'
    content += f'发生时间:' + time.strftime("%Y-%m-%d %H:%M:%S")
    if extra_content != '':
        content = content + '\n' + extra_content
    send_msg(Subject, content, agent_id)


def GetToken(Corpid, Secret):
    """获取access_token"""
    Url = "https://qyapi.weixin.qq.com/cgi-bin/gettoken"
    Data = {
        "corpid": Corpid,
        "corpsecret": Secret
    }
    r = requests.get(url=Url, params=Data, verify=False).json()
    c_time = int(time.time())
    if r.get('errcode') != 0:
        return False
    else:
        try:
            with open(token_file, 'r') as f:
                tokenlist = f.readlines()
                if len(tokenlist) == 2 and c_time < int(tokenlist[1].strip()):
                    token = tokenlist[0].strip()
                    expires = tokenlist[1].strip()
                else:
                    raise Exception
        except Exception:
            with open(token_file, 'w') as f:
                try:
                    token = r.get('access_token')
                    expires = int(r.get('expires_in', 0)) + c_time
                    f.write('%s\n%s' % (token, expires))
                except Exception as e:
                    token = ''
        return token


def SendMessage(Subject, Content, Agentid):
    """发送消息"""
    # 获取token信息
    Token = GetToken(Corpid, Secret)
    # 发送消息
    Url = "https://qyapi.weixin.qq.com/cgi-bin/message/send?access_token=%s" % Token
    Data = {
        "msgtype": "text",
        "agentid": Agentid,
        "text": {"content": Subject + '\n----------------------------------------------------------\n' + Content},
        "touser": '@all',
        "safe": "0"
    }
    r = requests.post(url=Url, data=json.dumps(Data), verify=False)
    # 如果发送失败，将重试三次
    n = 1
    while r.json()['errcode'] != 0 and n < 4:
        n = n + 1
        Token = GetToken(Corpid, Secret)
        if Token:
            Url = "https://qyapi.weixin.qq.com/cgi-bin/message/send?access_token=%s" % Token
            r = requests.post(url=Url, data=json.dumps(Data), verify=False)
    return r.json()['errcode']


def send_msg(Subject, Content, Agentid):
    for agentid in Agentid:
        Status = SendMessage(Subject, Content, agentid)


def init_event_status(db_conf=None):
    """ 初始化当天需要运行的sql,邮件事件状态event_status """
    try:
        with pymysql.connect(**db_conf) as conn:
            sql = f"""select id,mail_title from  mail_sql_conf
                                                  WHERE
                                                    -- 邮件事件是否被禁止
                                                    event_stop != 1
                                                    -- 当天的邮件未发送
                                                    AND event_status != 0
                                                    -- 邮件事件是否过期
                                                    AND
                                                    (sql_start_date <= date(now()) AND sql_end_date >= date(now()))
                                                    -- 当天邮件是否已经发送判断
                                                    AND
                                                    ((sql_cycle_type = 1
                                                        )
                                                    OR (sql_cycle_type = 2
                                                        AND date(now()) = DATE_ADD(DATE_SUB(CURDATE(), INTERVAL WEEKDAY(CURDATE()) DAY), INTERVAL sql_cycle_offset - 1 DAY))
                                                    OR (sql_cycle_type = 3
                                                        AND date(now()) = DATE_ADD(DATE_FORMAT(CURDATE(), '%Y-%m-01'), INTERVAL sql_cycle_offset - 1 DAY))
                                                    OR  (sql_cycle_type = 4
                                                        AND((sql_last_exec_date IS NOT NULL
                                                            AND DATE_FORMAT(CURDATE(), '%Y-%m-%d') = date_add( sql_last_exec_date, INTERVAL (datediff(DATE_FORMAT(CURDATE(), '%Y-%m-%d'), sql_last_exec_date) DIV sql_cycle_offset + 1 ) * sql_cycle_offset -1 DAY))
                                                            OR DATE_FORMAT(CURDATE(), '%Y-%m-%d') = date_add( sql_start_date, INTERVAL (datediff(DATE_FORMAT(CURDATE(), '%Y-%m-%d'), sql_start_date) DIV sql_cycle_offset + 1 ) * sql_cycle_offset -1 DAY))
                                                      ))"""
            cursor = conn.cursor()
            flag_init_sql = cursor.execute(sql)
            log_into_db('初始化当天sql', f'初始化邮件事件状态', db_conf=mail_config, event_type=0)
            if flag_init_sql:
                results = cursor.fetchall()
                for result in results:
                    _id, mail_title = result[0], result[1]
                    log_into_db(mail_title, '初始化邮件事件状态 set event_status = 0', db_conf=mail_config,
                                event_type=23)
                    mail_type_change(_id, mail_title=mail_title, db_conf=mail_config, event_type=0)
                    conn.commit()
    except pymysql.err.Error as e:
        log_into_db('初始化当天sql', f'初始化邮件事件状态出错 error: {str(e)}', db_conf=mail_config, event_type=23)
        wechat_alert(23, mail_title='初始化当天sql')


def mail_sql_backup(sql_text, db_conf=None, backup_table_name=[], mail_title=None, mail_description=None,
                    data_info=None, id=None):
    """mail_sql数据库备份

    :param id: 作为备份表备份那一条sql的标识，代表mail_sql 表中那一条sql的备份
    :param sql_text: sql语句
    :param db_conf: 数据库配置
    :param backup_table_name: 备份表名 list
    :param mail_title: 邮件标题-用于备份表
    :param mail_description: cursor.description
    :param data_info: 数据信息
    """
    try:
        tips_sql = ''
        backup_table_name_list = list()
        backup_table_name_list = backup_table_name.strip("[]").split(",")
        with pymysql.connect(**db_conf) as conn:
            for backup_table_name_one in backup_table_name_list:
                create_table_name = 'backup_' + id + backup_table_name_one + '_' + str(
                    time.strftime("%Y%m%d", time.localtime()))
                # 判斷create_table_name 是否存在
                sql_table_exist = f"select * from information_schema.tables where table_name = '{create_table_name}'"
                cursor = conn.cursor()
                cursor.execute(sql_table_exist)
                if cursor.rowcount == 0:
                    backup_sql = f"CREATE TABLE {create_table_name} ("
                    for field in mail_description:
                        field_name = field[0]
                        field_type = field[1]

                        # 根据字段类型确定相应的数据类型
                        if field_type == pymysql.STRING:
                            data_type = "VARCHAR(255)"
                        elif field_type == pymysql.NUMBER:
                            data_type = "INT"
                        elif field_type == pymysql.DATETIME:
                            data_type = "DATETIME"
                        else:
                            data_type = "VARCHAR(255)"  # 默认为 VARCHAR(255)

                        backup_sql += f"{field_name} {data_type}, "
                    backup_sql = backup_sql.rstrip(", ")
                    backup_sql += ")"
                    tips_sql = backup_sql
                    cursor.execute(backup_sql)
                    cursor.execute(f"ALTER TABLE {create_table_name} ADD  COLUMN dba_create_time DATETIME  ON UPDATE "
                                   f"CURRENT_TIMESTAMP;")
                    cursor.execute(f"ALTER TABLE {create_table_name} MODIFY COLUMN dba_create_time DATETIME DEFAULT "
                                   f"CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP;")
                    tips_sql = backup_sql
                column_list = ''
                for one_column_name in mail_description:
                    if one_column_name[0] != 'dba_create_time':
                        column_list += one_column_name[0] + ','
                column_list = column_list[:-1]
                for row in data_info:
                    values = '"' + '","'.join(str(i_column) for i_column in row) + '"'
                    values = values.replace('"None"', 'NULL')
                    # print(values)
                    backup_sql = f"insert into {create_table_name}({column_list}) values ({values}) "
                    print(backup_sql)
                    tips_sql = backup_sql
                    cursor.execute(backup_sql)
                conn.commit()
    except pymysql.err.Error as e:
        log_into_db('mail_sql_backup', f'备份mail_sql结果失败 sql: {tips_sql} error: {str(e)}', db_conf=mail_config,
                    event_type=24)
        wechat_alert(24, mail_title=mail_title)


if __name__ == '__main__':
    # 执行成功标识
    global flag_success
    flag_success = 0
    db_type = 'mysql'
    # Corpid是企业号的标识
    global Corpid
    Corpid = "ww980c7371d32a988e"
    # Secret是管理组凭证密钥
    global Secret
    Secret = "I2z7ecbDz7vmuhZefxMpHFnQ_ianZrZoqbxEeXlPje8"
    # agentid
    global agent_id
    agent_id = (1000007,)
    global token_file
    token_file = f'/tmp/qywechat_token.txt'
    # token_file = f'{os.getcwd()}\\qywechat_token.txt'

    # db_config = {
    #     'host': '172.29.28.193',
    #     'port': 3306,
    #     'database': 'sql_mail',
    #     'user': 'dzjroot',
    #     'password': '12345678'
    # }
    # db_config2 = {
    #     'host': '172.29.28.193',
    #     'port': 3307,
    #     'database': 'mysql',
    #     'user': 'dzjroot',
    #     'password': '123456'
    # }

    db_config2 = {
        'host': '172.30.70.45',
        'port': 3106,
        'database': 'dzjetl',
        'user': 'dzjetl',
        'password': 'EJhGcfpypgu0r0Xj'
    }

    db_config = {
        'host': '172.30.2.233',
        'port': 3306,
        'database': 'dzj',
        'user': 'shenxiang',
        'password': 'SHenXiang123'
    }

    # 邮件发送配置服务
    global mail_config
    # mail_config = {
    #     'host': '172.29.28.193',
    #     'port': 3306,
    #     'database': 'sql_mail',
    #     'user': 'dzjroot',
    #     'password': '12345678'
    # }
    mail_config = {
        'host': '172.30.70.11',
        'port': 3106,
        'database': 'sql_mail',
        'user': 'dzjetl',
        'password': 'bG01AUtBDJQvATvJ'
    }
    # mail_sql状态初始化
    if time.strftime("%H:%M") == '01:00':
        init_event_status(mail_config)
    # init_event_status(mail_config)
    # 查询 1.0.0 库
    # 获取需要执行的sql
    try:
        with pymysql.connect(**mail_config) as conn:
            cursor = conn.cursor()
            # 执行sql,获取当天需要执行的sql
            cursor.execute(f"""select sql_text,mail_title,mail_to,mail_cc,id,db_backup,backup_table_name,attachment_name from mail_sql_conf
                                                  WHERE
                                                    -- 邮件事件是否被禁止
                                                    event_stop != 1
                                                    -- 当天的邮件未发送
                                                    AND event_status = 0
                                                    -- 源数据库
                                                    AND db_source = 1
                                                    -- 邮件发送时间点判断
                                                    AND DATE_ADD(curtime(),INTERVAL 4 MINUTE) > send_time AND DATE_ADD(curtime(),INTERVAL -4 MINUTE) < send_time
                                                    -- 邮件事件是否过期
                                                    AND
                                                    (sql_start_date <= date(now()) AND sql_end_date >= date(now()))
                                                    -- 当天邮件是否已经发送判断
                                                    AND
                                                    ((sql_cycle_type = 1
                                                        )
                                                    OR (sql_cycle_type = 2
                                                        AND date(now()) = DATE_ADD(DATE_SUB(CURDATE(), INTERVAL WEEKDAY(CURDATE()) DAY), INTERVAL sql_cycle_offset - 1 DAY))
                                                    OR (sql_cycle_type = 3
                                                        AND date(now()) = DATE_ADD(DATE_FORMAT(CURDATE(), '%Y-%m-01'), INTERVAL sql_cycle_offset - 1 DAY))
                                                    OR  (sql_cycle_type = 4
                                                        AND((sql_last_exec_date IS NOT NULL
                                                            AND DATE_FORMAT(CURDATE(), '%Y-%m-%d') = date_add( sql_last_exec_date, INTERVAL (datediff(DATE_FORMAT(CURDATE(), '%Y-%m-%d'), sql_last_exec_date) DIV sql_cycle_offset + 1 ) * sql_cycle_offset -1 DAY))
                                                            OR DATE_FORMAT(CURDATE(), '%Y-%m-%d') = date_add( sql_start_date, INTERVAL (datediff(DATE_FORMAT(CURDATE(), '%Y-%m-%d'), sql_start_date) DIV sql_cycle_offset + 1 ) * sql_cycle_offset -1 DAY))
                                                      ))""")
            sql_results = cursor.fetchall()
            if cursor.rowcount != 0:
                # mail_name,mail_sql,mail_user,mail_cc,sub_title
                mail_info = [[mail_conf_row[1], mail_conf_row[0], mail_conf_row[2], mail_conf_row[3], mail_conf_row[7],
                              mail_conf_row[4], mail_conf_row[5], mail_conf_row[6]] for mail_conf_row in sql_results]
                sendmail = send_mail(db_type, db_config)
                sendmail.send_mail(mail_info)
                flag_success += 1
        # 查询 2.0 库
        with pymysql.connect(**mail_config) as conn:
            cursor = conn.cursor()
            # 执行sql,获取当天需要执行的sql
            cursor.execute(f"""select sql_text,mail_title,mail_to,mail_cc,id,db_backup,backup_table_name,attachment_name from mail_sql_conf
                                                  WHERE
                                                    -- 邮件事件是否被禁止
                                                    event_stop != 1
                                                    -- 当天的邮件未发送
                                                    AND event_status = 0
                                                    -- 源数据库
                                                    AND db_source = 2
                                                    -- 邮件事件是否过期
                                                    AND
                                                    (sql_start_date <= date(now()) AND sql_end_date >= date(now()))
                                                    -- 邮件发送时间点判断
                                                    AND DATE_ADD(curtime(),INTERVAL 4 MINUTE) > send_time AND DATE_ADD(curtime(),INTERVAL -4 MINUTE) < send_time
                                                    -- 当天邮件是否发送判断
                                                    AND
                                                    ((sql_cycle_type = 1
                                                        )
                                                    OR (sql_cycle_type = 2
                                                        AND date(now()) = DATE_ADD(DATE_SUB(CURDATE(), INTERVAL WEEKDAY(CURDATE()) DAY), INTERVAL sql_cycle_offset - 1 DAY))
                                                    OR (sql_cycle_type = 3
                                                        AND date(now()) = DATE_ADD(DATE_FORMAT(CURDATE(), '%Y-%m-01'), INTERVAL sql_cycle_offset - 1 DAY))
                                                    OR  (sql_cycle_type = 4
                                                        AND((sql_last_exec_date IS NOT NULL
                                                            AND DATE_FORMAT(CURDATE(), '%Y-%m-%d') = date_add( sql_last_exec_date, INTERVAL (datediff(DATE_FORMAT(CURDATE(), '%Y-%m-%d'), sql_last_exec_date) DIV sql_cycle_offset + 1 ) * sql_cycle_offset -1 DAY))
                                                            OR DATE_FORMAT(CURDATE(), '%Y-%m-%d') = date_add( sql_start_date, INTERVAL (datediff(DATE_FORMAT(CURDATE(), '%Y-%m-%d'), sql_start_date) DIV sql_cycle_offset + 1 ) * sql_cycle_offset -1 DAY))
                                                      ))""")
            sql_results = cursor.fetchall()
            if cursor.rowcount != 0:
                # mail_name,mail_sql,mail_user,mail_cc,sub_title
                mail_info2 = [[mail_conf_row[1], mail_conf_row[0], mail_conf_row[2], mail_conf_row[3], mail_conf_row[7],
                               mail_conf_row[4], mail_conf_row[5], mail_conf_row[6]] for mail_conf_row in sql_results]
                for mail_conf_row in sql_results:
                    mail_type_change(mail_conf_row[4], mail_title=mail_conf_row[1], db_conf=mail_config, event_type=1)
                sendmail2 = send_mail(db_type, db_config2)
                sendmail2.send_mail(mail_info2)
                flag_success += 1
    except pymysql.err.Error as e:
        log_into_db('没有mail_sql需要执行', f'沒有获取到mail_sql , error: {str(e)}', db_conf=mail_config, event_type=11)
        wechat_alert(11)

    log_into_db(f'邮件事件结束', '', db_conf=mail_config, event_type=50)
    if flag_success != 0:
        wechat_alert(4)
