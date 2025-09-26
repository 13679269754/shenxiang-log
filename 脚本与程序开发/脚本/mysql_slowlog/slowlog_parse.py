import configparser
import os
import datetime
import hashlib
import re
import sys
import time
import json
import socket
import platform
from itertools import islice
from prettytable import PrettyTable
import pandas as pd
import logging as logging
import duckdb
from jinja2 import Environment, FileSystemLoader, Template
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.header import Header
import smtplib
from openpyxl import Workbook
from openpyxl.styles import PatternFill, colors, Font, Alignment
from elasticsearch import Elasticsearch


import warnings
warnings.filterwarnings("ignore")


pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)
pd.set_option('max_colwidth', 200)
pd.set_option('display.width',1000)


config_name = 'db_config.ini'
exepath = os.path.split(os.path.realpath(__file__))[0]
config_path = os.path.join(exepath, config_name)

config = configparser.ConfigParser()
config.read(config_path, encoding='utf-8-sig')

time_start_1 = time.time()
time_start = datetime.datetime.now()
time_day_format = time_start.strftime("%Y%m%d")
time_second_format = time_start.strftime("%Y%m%d_%H_%M_%S")

class ReadConfig:
    @staticmethod
    def __get_info(section,name):
        value = 0
        try:
            value = config.get(section, name)  # 通过config.get拿到配置文件中DATABASE的name的对应值
        except Exception as e:
            print(e)
            value = -1
        return value

    @staticmethod
    def get_path(name):
        return ReadConfig.__get_info('path',name)

    @staticmethod
    def get_elastic_connect(name):
        return ReadConfig.__get_info('elastic_connect', name)

    @staticmethod
    def get_parameter(name):
        return ReadConfig.__get_info('parameter', name)

def get_ip_address():
    # 创建一个socket对象
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        # 连接到一个公共的IP地址和端口
        s.connect(('8.8.8.8', 80))
        # 获取本地IP地址
        ip_address = s.getsockname()[0]
    except socket.error:
        ip_address = '127.0.0.1'
    finally:
        s.close()
    return ip_address
ipaddress = get_ip_address()


logger = ''
# 定义自定义日志级别的数值
SUCCESS_LEVEL = 25

# 定义自定义日志级别的日志方法
def success(self, message, *args, **kwargs):
    if self.isEnabledFor(SUCCESS_LEVEL):
        self._log(SUCCESS_LEVEL, message, args, **kwargs)

def set_logger(log_dir):
    log_file = os.path.join(log_dir, f'slowlog_parse_report.log')
    global logger
    # 定义自定义日志级别的名称
    logging.addLevelName(SUCCESS_LEVEL, "SUCCESS")
    # 将自定义日志级别方法添加到logger对象
    logging.Logger.success = success

    # 定义日志格式
    try:
        import colorlog
        # 创建彩色日志格式
        formatter = colorlog.ColoredFormatter(
            '%(log_color)s%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S',
            log_colors={
                'DEBUG': 'cyan',
                'INFO': 'white',
                'WARNING': 'yellow',
                'ERROR': 'red',
                'CRITICAL': 'red,bg_white',
                'SUCCESS':'green'
            }
        )
    except:
        formatter = logging.Formatter('%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s',datefmt='%Y-%m-%d %H:%M:%S')

    # 创建Logger对象
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)

    # 创建文件处理器
    file_handler = logging.FileHandler(log_file)
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(formatter)

    # 创建终端处理器
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.DEBUG)
    console_handler.setFormatter(formatter)

    # 将处理器添加到Logger对象
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

def print_info():
    title_name = 'MySQL Slow Log Parse Tool'
    k = PrettyTable(field_names=[title_name])
    k.align[title_name] = "l"  # 以name字段左对齐
    k.padding_width = 1  # 填充宽度
    k.add_row(["Tool Version: 1.0"])
    print(k.get_string(sortby=title_name, reversesort=False))

def read_db_to_xlsx(sql_result_list, sql_result_list2, excelname):
    # 循环数据写入工作表1内容
    jb_date_lists = sql_result_list
    jb_date_list = jb_date_lists[0]
    descripte = jb_date_lists[1]
    excel_rows_num = jb_date_lists[2]
    # 写入工作表2内容
    jb_date_lists2 = sql_result_list2
    jb_date_list2 = jb_date_lists2[0]
    descripte2 = jb_date_lists2[1]
    excel_rows_num2 = jb_date_lists2[2]
    # 要创建的xlsx名称
    dest_filename = excelname
    # 新建工作簿
    wb = Workbook()
    # 新建工作表1
    ws1 = wb.active
    ws1.title = "最大执行耗时排序"
    # 工作表1列名
    for i in range(0, len(descripte)):
        ws1.cell(row=1, column=i + 1, value=descripte[i])
    # 工作表1写入数据
    for i in range(2, len(jb_date_list) + 2):
        for j in range(0, len(descripte)):
            if jb_date_list[i - 2][j] is None:
                ws1.cell(row=i, column=j + 1, value='')
            else:
                ws1.cell(row=i, column=j + 1, value=jb_date_list[i - 2][j])
    # 工作表1设置单元格背景
    fill_a1 = PatternFill("solid", fgColor="1874CD")
    ws1["A1"].fill = fill_a1
    fill_b1 = PatternFill("solid", fgColor="1874CD")
    ws1["B1"].fill = fill_b1
    fill_c1 = PatternFill("solid", fgColor="1874CD")
    ws1["C1"].fill = fill_c1
    fill_d1 = PatternFill("solid", fgColor="1874CD")
    ws1["D1"].fill = fill_d1
    fill_e1 = PatternFill("solid", fgColor="1874CD")
    ws1["E1"].fill = fill_e1
    fill_f1 = PatternFill("solid", fgColor="1874CD")
    ws1["F1"].fill = fill_f1
    fill_g1 = PatternFill("solid", fgColor="1874CD")
    ws1["G1"].fill = fill_g1
    fill_h1 = PatternFill("solid", fgColor="1874CD")
    ws1["H1"].fill = fill_h1
    fill_i1 = PatternFill("solid", fgColor="1874CD")
    ws1["I1"].fill = fill_i1
    fill_j1 = PatternFill("solid", fgColor="1874CD")
    ws1["J1"].fill = fill_j1
    # 工作表1设置单元格文字颜色
    font_set = Font(color=colors.WHITE, bold=True)
    ws1['A1'].font = font_set
    ws1['B1'].font = font_set
    ws1['C1'].font = font_set
    ws1['D1'].font = font_set
    ws1['E1'].font = font_set
    ws1['F1'].font = font_set
    ws1['G1'].font = font_set
    ws1['H1'].font = font_set
    ws1['I1'].font = font_set
    ws1['J1'].font = font_set
    # 工作表1设置列宽
    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 100
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 20
    ws1.column_dimensions['E'].width = 25
    ws1.column_dimensions['F'].width = 20
    ws1.column_dimensions['G'].width = 10
    ws1.column_dimensions['H'].width = 10
    ws1.column_dimensions['I'].width = 10
    ws1.column_dimensions['J'].width = 20
    # 工作表1设置单元格对齐,替换参数的列为填充fill避免文本太长全部显示出来
    ws_area = ws1["I2:I%s" % excel_rows_num]
    for i in ws_area:
        for j in i:
            j.alignment = Alignment(horizontal='fill', vertical='center', wrap_text=False)
    # 工作表1其余列为左对齐
    ws_area = ws1["A2:H%s" % excel_rows_num]
    for i in ws_area:
        for j in i:
            j.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    # 工作表1冻结首行，方便下拉的时候能一直显示列名,设置A1没效果，所以要设置为A2即A2之上的第一行冻结
    ws1.freeze_panes = 'A2'
    # 以上是工作表1完毕
    # 以下是新建工作表2
    ws2 = wb.create_sheet("最多执行次数排序")
    # 工作表2列名
    for i in range(0, len(descripte2)):
        ws2.cell(row=1, column=i + 1, value=descripte2[i])
    # 工作表2写入数据
    for i in range(2, len(jb_date_list2) + 2):
        for j in range(0, len(descripte2)):
            if jb_date_list2[i - 2][j] is None:
                ws2.cell(row=i, column=j + 1, value='')
            else:
                ws2.cell(row=i, column=j + 1, value=jb_date_list2[i - 2][j])
    # 设置工作表2的背景
    ws2["A1"].fill = fill_a1
    ws2["B1"].fill = fill_b1
    ws2["C1"].fill = fill_c1
    ws2["D1"].fill = fill_d1
    ws2["E1"].fill = fill_e1
    ws2["F1"].fill = fill_f1
    ws2["G1"].fill = fill_g1
    ws2["H1"].fill = fill_h1
    ws2["I1"].fill = fill_i1
    ws2["J1"].fill = fill_j1
    # 设置工作表2的文字颜色
    ws2['A1'].font = font_set
    ws2['B1'].font = font_set
    ws2['C1'].font = font_set
    ws2['D1'].font = font_set
    ws2['E1'].font = font_set
    ws2['F1'].font = font_set
    ws2['G1'].font = font_set
    ws2['H1'].font = font_set
    ws2['I1'].font = font_set
    ws2['J1'].font = font_set
    # 设置工作表2的列宽
    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 100
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 20
    ws2.column_dimensions['E'].width = 25
    ws2.column_dimensions['F'].width = 20
    ws2.column_dimensions['G'].width = 10
    ws2.column_dimensions['H'].width = 10
    ws2.column_dimensions['I'].width = 10
    ws2.column_dimensions['J'].width = 20
    # 工作表2设置单元格对齐,替换参数的列为填充fill避免文本太长全部显示出来
    ws_area = ws2["I2:I%s" % excel_rows_num2]
    for i in ws_area:
        for j in i:
            j.alignment = Alignment(horizontal='fill', vertical='center', wrap_text=False)
    # 工作表2,其余列为左对齐
    ws_area = ws2["A2:H%s" % excel_rows_num2]
    for i in ws_area:
        for j in i:
            j.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    # 工作表2,冻结首行，方便下拉的时候能一直显示列名,设置A1没效果，所以要设置为A2即A2之上的第一行冻结
    ws2.freeze_panes = 'A2'
    # 创建xlsx
    wb.save(filename=dest_filename)

class slowlog(object):
    def __init__(self):
        ini_file = ReadConfig()
        # 1 文件格式解析 2 数据库格式解析
        self.parse_type = int(ini_file.get_parameter('parse_type'))
        self.read_size = int(ini_file.get_parameter('read_size'))
        self.log_read_size = int(ini_file.get_parameter('log_read_size'))
        self.log_parse_size = int(ini_file.get_parameter('log_parse_size'))
        self.slow_log_path = ini_file.get_path('slow_log_path')
        if ini_file.get_path('slow_log_name') and ini_file.get_path('slow_log_name') != '[]':
            self.slow_log_name = json.loads(ini_file.get_path('slow_log_name'))
        else:
            self.slow_log_name = ''
        self.slow_result  = []

        self.hosts = ini_file.get_elastic_connect('hosts').split(',')
        self.user = ini_file.get_elastic_connect('user')
        self.password = ini_file.get_elastic_connect('passwd')
        self.slow_log_index = ini_file.get_elastic_connect('slow_log_index')

        self.time_type = int(ini_file.get_parameter('time_type'))

        now = datetime.datetime.now()

        if self.time_type == 1:
            time1 = now.strftime('%Y-%m-%d %H:%M:%S') if ini_file.get_parameter('start_time') == 'now' else ini_file.get_parameter('start_time')
            time2 = now.strftime('%Y-%m-%d %H:%M:%S') if ini_file.get_parameter('end_time') == 'now' else ini_file.get_parameter('end_time')
            self.query_start_time = datetime.datetime.strptime(time1,'%Y-%m-%d %H:%M:%S')
            self.query_end_time = datetime.datetime.strptime(time2,'%Y-%m-%d %H:%M:%S')
        elif self.time_type == 2:
            delay_day = now - datetime.timedelta(days=int(ini_file.get_parameter('time_delay')))
            delay_day = delay_day.replace(hour=0, minute=0, second=0, microsecond=0)
            now_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
            self.query_start_time = delay_day
            self.query_end_time = now_day
        else:
            self.query_start_time = ''
            self.query_end_time = ''

        logger.warning(f'日志起始时间：{self.query_start_time}')
        logger.warning(f'日志结束时间：{self.query_end_time}')


    def run(self):
        if self.parse_type == 1:
            self.read_log_path()
        elif self.parse_type == 2:
            self.get_data_from_elasticsearch()
        df = self.create_df_data()
        return df

    def get_es_search_result(self, index, scroll='5m', timeout='1m', query = ''):
        client = Elasticsearch(self.hosts, basic_auth=(self.user, self.password), ssl_assert_hostname=False, verify_certs=False)
        if query:
            query = query
        else:
            query = {
                "match_all": {}
            }

        read_size = self.read_size
        queryData = client.search(
            index=index,
            scroll=scroll,
            timeout=timeout,
            size=read_size,
            query=query
        )

        mdata = queryData.get("hits").get("hits")
        count = len(mdata)

        if not mdata:
            print('empty')

        scroll_id = queryData["_scroll_id"]
        total = queryData["hits"]["total"]['value']

        for i in range(int(total / read_size) + 1):
            res = client.scroll(scroll_id=scroll_id, scroll='5m')
            mdata = mdata + res["hits"]["hits"]
            count += len(res["hits"]["hits"])
            logger.info(f'已获取行数：{count}')

        return mdata


    def get_data_from_elasticsearch(self):
        query = {
            "range": {
                    "@timestamp": {
                        "format": "yyyy-MM-dd HH:mm:ss"
                    }
                }
        }

        if self.query_start_time:
            query['range']['@timestamp']['gte'] = f"{self.query_start_time.strftime('%Y-%m-%d %H:%M:%S')}"

        if self.query_start_time:
            query['range']['@timestamp']['lte'] = f"{self.query_end_time.strftime('%Y-%m-%d %H:%M:%S')}"

        data = self.get_es_search_result(self.slow_log_index,query = query)

        for i in data:
            new_dic = {}
            info = i.get('_source')
            d = info.get('@timestamp')  # 2024-05-28T01:32:59.144Z
            _date = datetime.datetime.strptime(d, "%Y-%m-%dT%H:%M:%S.%fZ")
            local_time = _date + datetime.timedelta(hours=8)
            end_time = local_time
            new_dic['ipaddress'] = info.get('agent').get('name')
            new_dic['datetime_value'] = end_time
            new_dic['query_time'] = info.get('query_time')
            new_dic['lock_time'] = info.get('lock_time')
            new_dic['rows_sent'] = info.get('rows_sent')
            new_dic['rows_examined'] = info.get('rows_examined')
            new_dic['client_ip'] = info.get('clientip')
            new_dic['user'] = info.get('user')
            new_dic['query'] = info.get('sql_info')
            new_dic['timestamp'] = info.get('sql_time')
            self.slow_result.append(new_dic)


    def read_log_path(self):

        if self.slow_log_path:
            file_list = []
            if self.slow_log_name:
                for name in self.slow_log_name:
                    file_list.append(os.path.join(self.slow_log_path,name))
            else:
                # 获取慢日志目录下的所有文件名
                for root, dirs, files in os.walk(self.slow_log_path):
                    for name in files:
                        file_list.append(os.path.join(root, name))

            for file_name in file_list:
                logger.info(f"开始分析慢日志文件：{file_name}")
                # 开始使用read_log_lines分析日志每一行
                self.read_log_lines(file_name)
            logger.info(f'总共生成 {len(self.slow_result)} 条慢日志')
        else:
            logger.error(f'获取慢日志文件路径 {self.slow_log_path} 异常，请检查配置文件！')
            return 0
        return 1

    def line_parse(self,line,log,first_log=True,database_info = ''):
        line = line.strip()
        if line[0:7] == "# Time:":
            if first_log:
                first_log = False
            else:
                self.slow_result.append(log)
                log = self.add_log()
            log['datetime'] = line
            log['database'] = database_info
            log['datetime_value'] = datetime.datetime(int(line[8:12]), int(line[13:15]), int(line[16:18]), int(line[19:21]), int(line[22:24]), int(line[25:27]))
        elif line[0:12] == "# User@Host:":
            log['database_host'] = line
            host = line.split()
            log['user'] = host[2].split('[')[0]
            log['client_ip'] = host[4].replace('[','').replace(']','')
        elif line[0:13] == "# Query_time:":
            log['time'] = line
            # Get number values for query_time, lock_time, rows_sent, rows_examined
            time = line.split()
            log['query_time'] = float(time[2])
            log['lock_time'] = float(time[4])
            log['rows_sent'] = float(time[6])
            log['rows_examined'] = float(time[8])
        elif line[0:3] == "use":
            log['database'] = line
            database_info = line
        elif line[0:14] == "SET timestamp=":
            log['timestamp'] = line
        else:
            log['query'] = log['query'] + ' ' +  line
        return first_log,log,database_info

    def add_log(self):
        log = {
            "database":"",
            "database_host":"",
            "time":"",
            "timestamp":"",
            "statement":"",
            "query":"",
            "client_ip":"",
            "user":"",
            "query_time":0,
            "lock_time":0,
            "rows_sent":0,
            "rows_examined":0,
            "datetime_value":datetime.time(0, 0)
        }
        return log
    def read_log_lines(self, filename):
        logger.info(f'开始读取文件 {filename} 每次读取 {self.read_size} 行')
        read_count = 0  # 慢日志读取的行数
        with open(filename, "r", encoding='utf-8', errors='ignore') as f:
            first_log = True
            database_info = ''
            log = self.add_log()
            lines = list(islice(f, 3, self.read_size))
            while True:
                if not lines:
                    break
                for line in lines:
                    read_count += 1
                    first_log, log, database_info = self.line_parse(line, log, first_log, database_info)

                    if read_count % self.log_read_size == 0:
                        logger.info(f'{filename} 已处理行数：{read_count}')
                lines = list(islice(f, 0, self.read_size))


        self.slow_result.append(log)
        logger.info(f'{filename} 总共处理行数：{read_count}')
        logger.success(f'{filename} 读取文件完毕')



    def create_df_data(self,ipaddress = ipaddress):
        list_parse_log = []
        logger.info(f'开始正则解析日志！')
        start_time = time.time()
        count = 0

        pattern_list=[
            [r' +',' '],            # 多个空格格式化为1个空格
            [r'/\*.*?\*/\s',""],    # 去除注释，即/* */的注释内容
            [r'=\s*\'\'',"= ?"],
            [r'=\s*""',"= ?"],
            [r'\'\'',""],           # 去掉SQL语句条件文本内容里，包含的单引号或者双引号
            [r'\'.*?\'',"?"],       # 把单引号包围的字符串内容连同单引号一起替换为问号
            [r'".*?"', "?"],        # 把双引号包围的字符串内容连同双引号一起替换为问号
            [r'\bfalse\b|\btrue\b', "?"],   # 不区分大小写的true或者false替换为问号
            [r'=\s*[0-9+-][0-9a-f.xb+-]*', "= ?"],  # where条件里的数字替换为问号
            [r'=\s*\b[0-9+-][0-9a-f.xb+-]*', "= ?"],  # where条件里的数字替换为问号
            [r'\bis\s+null\b', "= ?"],
            [r'\bis\s+not\s+null\b', "!= ?"],
            [r'=\s*null', "= ?"],          # where里的null替换为问号
            [r'<>\s*\bnull\b', "!= ?"],    # where里的null替换为问号
            [r'values\s+[\(.*\)\s*]+', "values(?+)"],
            [r'in\s*\(.*?\)', "in(?+)"],
            [r'\b(in|values?)(?:[\s,]*\([\s?,]*\))+', r'\1(?+)'],  # 将in,values匹配的多个问号替换为"?+"
            [r'\blimit\s+\d+\s*offset\s\d+', ' limit ?'],  # limit offset 统一格式化为limit ?
            [r'\blimit\s+\d+', ' limit ?'],  # limit offset 统一格式化为limit ?
        ]


        compile_pattern_list = []
        for line in pattern_list:
            pattern = re.compile(line[0],flags=re.I)
            compile_pattern_list.append([pattern,line[1]])


        for slow_item in self.slow_result:
            if self.parse_type == 2:
                ipaddress = slow_item['ipaddress']
                exec_client = ''
            else:
                ipaddress = ipaddress
                exec_client = slow_item['database_host']
            sql_finish_time = slow_item['datetime_value'].strftime('%Y-%m-%d %H:%M:%S')

            if self.query_start_time:
                if slow_item['datetime_value'] < self.query_start_time:
                    continue

            if self.query_end_time:
                if slow_item['datetime_value'] > self.query_end_time:
                    continue


            query_time = slow_item['query_time']
            lock_time = slow_item['lock_time']
            rows_sent = slow_item['rows_sent']
            rows_examined = slow_item['rows_examined']
            user = slow_item['user']
            schema = slow_item['database'].replace('use ','').replace(';','') if slow_item.get('database') else ''
            client_ip = slow_item['client_ip']
            sql_text = slow_item['query'].replace('\n', ' ').strip()

            sql_start_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(int(slow_item['timestamp'].replace('SET timestamp=', '').replace(';', ''))))


            if re.match(r'/\*.*\*/\s', sql_text):  # 先检测是否包含注释行/* */
                str_len = len(re.findall(r'/\*.*\*/\s', sql_text)[0])  # 获取注释行开头字符串的长度
                optype = re.split(r'\s+', sql_text[str_len:])[0].upper()  # 截取SQL操作类型
            else:
                optype = re.split(r'\s+', sql_text)[0].upper()  # 空格作为分割截取SQL操作类型

            query = sql_text.strip().lower()  # 去除首尾多余的空格以及字符串转为小写
            # query = sql_text.rstrip(';')  # 去除结尾的分号
            for line in compile_pattern_list:
                query = line[0].sub(line[1], query)

            if re.match(r"\b(select\s.*?)(?:(\sunion(?:\sall)?)\s)+", query, flags=re.I):  # 最后一个union all使用repeat注释
                query = re.sub(r"\b(select\s.*?)(?:(\sunion(?:\sall)?)\s\1)+", r'\1 /*repeat\2*/', query, flags=re.I)
            # sql文本如果带有order by asc条件，去掉order by里的asc这个冗余的字符
            if re.match(r'(.*) order\s+by (.*?)\s+asc', query, flags=re.I):
                query = re.sub(r'\s+ASC', '', query, flags=re.I)

            query = re.sub(r"\A\s*call\s+(\S+)\(.*?\).*;", r'call \1', query, flags=re.I)  # call形式的存储过程，去掉括号
            # print(query)  # 去掉参数之后格式化的sql形式
            md5_query = hashlib.md5(query.encode()).hexdigest().upper()  # 计算文本字符串转为md5的十六进制字符串
            # print(md5_query)
            # 以下将慢日志各信息存到list。以便组成dataframe
            list_parse_log.append([ipaddress, sql_finish_time, sql_start_time, exec_client, client_ip, user,schema,query_time, lock_time, rows_sent, rows_examined, sql_text, query, optype,md5_query])
            count += 1
            program_time = time.time()
            # 以下为慢日志分析部分
            if count % self.log_parse_size == 0:
                logger.info(f"""已处理 {count} 条日志，耗时{'{:.2f}'.format(program_time - start_time)}秒""")
        df = pd.DataFrame(list_parse_log, columns=['ipaddress','sql_finish_time','sql_start_time','exec_client', 'client_ip', 'user','schema', 'query_time', 'lock_time', 'rows_sent', 'rows_examined', 'sql_text', 'query', 'optype', 'md5_query'])


        end_time = time.time()
        logger.success(f"""解析日志完毕，总共处理 {count} 条日志，耗时 {'{:.2f}'.format(end_time - start_time)} 秒！""")
        return df



class query_report():
    def __init__(self,df,report_dir):
        ini_file = ReadConfig()
        self.slow_log_path = ini_file.get_path('slow_log_path')
        self.slow_log_name = ini_file.get_path('slow_log_name')
        self.query_total_size = int(ini_file.get_parameter('query_total_size'))
        self.parse_type = int(ini_file.get_parameter('parse_type'))
        self.report_type = ini_file.get_parameter('report_type').split(',')
        self.if_debug = int(ini_file.get_parameter('if_debug'))
        self.debug_print_report = int(ini_file.get_parameter('debug_print_report'))
        self.debug_print_df = int(ini_file.get_parameter('debug_print_df'))
        self.if_sendmail = int(ini_file.get_parameter('if_sendmail'))
        self.if_sendwx = int(ini_file.get_parameter('if_sendwx'))
        self.mail_name_sub = ini_file.get_parameter('mail_name_sub')
        self.mail_user = ini_file.get_parameter('mail_user')
        self.mail_cc = ini_file.get_parameter('mail_cc')


        if int(ini_file.get_parameter('if_data')) == 1:
            self.if_data = True
        else:
            self.if_data = False


        self.query_total_size_sql = ''
        if self.query_total_size != 0:
            self.query_total_size_sql = f'limit {self.query_total_size}'

        self.df = df
        if self.if_debug == 1 and self.debug_print_df == 1:
            print(self.df)

        self.report_dir = report_dir

        self.excel_file = f"slow_sql_report_{time_second_format}.xlsx"
        self.html_file = f"slow_sql_report_{time_second_format}.html"
        self.txt_file = f"slow_sql_report_{time_second_format}.txt"

        new_df = self.df
        unqiue_sql = "select distinct ipaddress from new_df"
        df_sql = duckdb.sql(unqiue_sql).df()
        self.ipaddress_data = [row['ipaddress'] for v_all_data, row in df_sql.iterrows()]  # 每一行转为元组tuple，所有数据以列表list输出

    def send_mail(self, mail_info, message_html_info, file_path):



        mail_name = mail_info['mail_name']
        mail_user = mail_info['mail_user']
        mail_cc = mail_info['mail_cc']

        msg = message_html_info
        mime = MIMEMultipart()
        mime.attach(MIMEText(msg, 'html', 'utf-8'))
        for i in file_path:
            att = MIMEApplication(open(i, 'rb').read())
            _, extension = os.path.splitext(i)
            suffix = extension[1:]
            att.add_header('Content-Disposition', 'attachment', filename=Header(f'{mail_name}.{suffix}', 'utf-8').encode())
            mime.attach(att)

        mime['To'] = mail_user
        mime['Cc'] = mail_cc
        mime['Subject'] = mail_name
        mime['From'] = 'DBA@dazhuanjia.com'
        mail_user = mail_user.split(',')

        if mail_cc:
            mail_user.extend(mail_cc.split(','))

        send = smtplib.SMTP("localhost")
        send.sendmail('localhost', mail_user, mime.as_string())
    def send_wx(self,file_name,content):
        from send_wx_text import send
        send(file_name,content)

    def run(self):
        file_path = []
        if '1' in self.report_type:
            txt_file_name = self.query_report_to_txt()
            if os.path.exists(txt_file_name):
                file_path.append(txt_file_name)
            if self.if_debug == 1 and self.debug_print_report == 1:
                file_name = os.path.join(self.report_dir, self.txt_file)
                with open(file_name, 'r') as f:
                    for i in f.readlines():
                        print(i.strip())

        if '2' in self.report_type:
            excel_file = os.path.join(self.report_dir, self.excel_file)

            # 生成按执行时间维度统计结果
            sql_result_list = self.query_max_exec_time()
            logger.success(f'query_max_exec_time finish')

            # 生成按执行次数分析的统计结果
            sql_result_list2 = self.query_max_exec_count()
            logger.success(f'query_max_exec_count finish')

            if len(sql_result_list[0]) > 0 and len(sql_result_list2[0]) > 0:
                # 将以上2个结果写入到同一个excel不同的sheet
                logger.success(f'正在生成Excel')
                read_db_to_xlsx(sql_result_list, sql_result_list2, excel_file)
                logger.success(f'生成Excel完毕')

            else:
                logger.error(f'没有查询到慢日志数据，生成Excel以及散点图失败，请检查日志文件！')

            if os.path.exists(excel_file):
                file_path.append(excel_file)

        if '3' in self.report_type:
            html_file_name,html_content = self.query_report_to_html()

            if os.path.exists(html_file_name):
                file_path.append(html_file_name)
        logger.success(f"解析慢日志完毕，请查看慢日志报告 报告目录：{self.report_dir}", )


        if self.if_sendmail == 1:
            mail_info = {
                'mail_name': f'{time_day_format}-{ipaddress} {self.mail_name_sub}',
                'mail_user':self.mail_user,
                'mail_cc':self.mail_cc,
            }
            self.send_mail(mail_info,f'数据查询完成，请查收.\n',file_path)

            logger.success(f"发送邮件成功，请查收", )
        if self.if_sendwx == 1:
            content_info = f'{time_day_format}-{ipaddress} {self.mail_name_sub}'
            for file_name in file_path:
                self.send_wx(file_name,content_info)
            logger.success(f"发送微信成功，请查收", )


    def query_report_to_txt(self):

        file_name = os.path.join(self.report_dir, self.txt_file)
        logger.info(f'开始生成 TXT 慢日志报告: {file_name}')

        with open(file_name, 'a+', encoding='utf-8', errors='ignore') as rpt:
            self.txt_query_report(rpt)

            if len(self.ipaddress_data) > 1:
                for i in self.ipaddress_data:
                    self.txt_query_report(rpt, ipaddress = i)
                    self.txt_query_detail_report(rpt, ipaddress = i)
            else:
                self.txt_query_detail_report(rpt)
        return file_name
    def query_report_to_html(self):
        file_name = os.path.join(self.report_dir, self.html_file)
        logger.info(f'开始生成 HTML 慢日志报告: {file_name}')

        env = Environment(loader=FileSystemLoader(os.path.dirname(__file__)))
        template = env.get_template('slowlog_template.html')

        total_info ,profile_info = self.html_query_report(template)

        instance_info = {}
        detail_info = {}
        query_id_list = []
        if self.query_total_size > 0:
            for k,v in profile_info.items():
                for i in v['data']:
                    query_id_list.append(i.get('query_id'))
        query_id_list = list(set(query_id_list))

        if len(self.ipaddress_data) > 1:
            for i in self.ipaddress_data:
                instance_info[i] = {}
                instance_info[i]['total_info'],instance_info[i]['profile_info'] = self.html_query_report(template, ipaddress=i)
                detail_info[i] = self.html_query_detail_report(template, ipaddress=i,query_id_list=query_id_list)
        else:
            detail_info['total'] = self.html_query_detail_report(template,query_id_list=query_id_list)

        html_content = template.render(total_info=total_info, profile_info=profile_info,instance_info=instance_info, detail_info = detail_info)

        with open(file_name, 'w') as f:
             f.write(html_content)

        return file_name, html_content

    def html_query_detail_report(self, html_template, ipaddress = '',query_id_list=[]):
        logger.info(f'开始生成{ipaddress} HTML 慢日志详情')

        # 以下是每个SQL的详细情况
        detail_info  = self.query_detail_info(ipaddress,query_id_list)

        if ipaddress != '':
            title = f'# {ipaddress} Parse Detail Report'
        else:
            title = f'# Total Parse Detail Report'

        return_info = {'title':title.replace('#','').strip(),'data':[]}
        for i, row in detail_info.iterrows():
            info = self.write_query_detail(row, '', 'html')
            return_info['data'].append(info)

        return return_info
    def html_query_report(self, html_template, ipaddress = ''):

        logger.info(f'开始生成{ipaddress}慢日志 HTML 汇总报告')

        total_info = {}
        if ipaddress != '':
            title = f'{ipaddress} Parse Report'
        else:
            title = f'Total Parse Report'

        total_info['title'] = title

        total_info['current_date'] = str(datetime.datetime.now())
        total_info['hostname'] = str(platform.node())
        total_info['ipaddress'] = str(get_ip_address()) if ipaddress == '' else ipaddress
        total_info['file_dir'] = self.slow_log_path if self.parse_type == 1 else ''
        total_info['file_name'] = self.slow_log_name if self.parse_type == 1 else ''
        total_info['data_from'] = 'Get Data From ElasticSearch' if self.parse_type == 2 else 'Get Data From Slow Log File'

        total_info_dict = self.query_total_info(ipaddress)
        total_info['overall'] = f"{total_info_dict['total']} total, {str(total_info_dict['unique'])} unique, {str(total_info_dict['qps'])} QPS, {str(total_info_dict['concurrency'])} concurrency"
        total_info['time_range'] = f"{total_info_dict['min_time']} to {total_info_dict['max_time']}"
        total_info['data'] = total_info_dict['total_data']


        profile_info ={}
        data = self.query_profile('query_time', 'query_id',ipaddress)
        data_list = self.write_query_profile(data,rpt = '',write_type = 'html')
        profile_info['Query Time Sort Profile(By QUERY_ID)'] = {'data': data_list, 'html_model_target': 1}

        if self.if_data:
            data = self.query_profile('query_time', 'sql_text',ipaddress)
            data_list = self.write_query_profile(data, rpt='', write_type='html')
            profile_info['Query Time Sort Profile(By SQL_TEXT)'] = {'data': data_list, 'html_model_target': 2}

        data = self.query_profile('query_count', 'query_id',ipaddress)
        data_list = self.write_query_profile(data, rpt='', write_type='html')
        profile_info['Calls Sort Profile(By QUERY_ID)']= {'data': data_list, 'html_model_target': 3}

        if self.if_data:
            data = self.query_profile('query_count', 'sql_text',ipaddress)
            data_list = self.write_query_profile(data, rpt='', write_type='html')
            profile_info['Calls Sort Profile(By SQL_TEXT)']= {'data': data_list, 'html_model_target': 4}

        data = self.query_profile('lock_time', 'query_id',ipaddress)
        data_list = self.write_query_profile(data, rpt='', write_type='html')
        profile_info['Lock Time Sort Profile(By QUERY_ID)']= {'data': data_list, 'html_model_target': 5}

        if self.if_data:
            data = self.query_profile('lock_time', 'sql_text',ipaddress)
            data_list = self.write_query_profile(data, rpt='', write_type='html')
            profile_info['Lock Time Sort Profile(By SQL_TEXT)']= {'data': data_list, 'html_model_target': 6}

        logger.success(f'生成{ipaddress}慢日志汇总报告')
        return total_info,profile_info


    def get_df_from_ipaddress(self, ipaddress = ''):
        if ipaddress:
            df = self.df[self.df['ipaddress'] == ipaddress]
        else:
            df = self.df
        return df

    def num_format(self,info,type = ''):
        try:
            if pd.isnull(info):
                return ''
            if type == '' or type.find('time') > 0:
                if info < 0.001:
                    return '{:.2f}'.format(info * 1000000) + ' us'
                elif info < 1 and info >= 0.001:
                    return '{:.2f}'.format(info * 1000) + ' ms'
                elif info < 1800 and info >= 1:
                    return '{:.2f}'.format(info) + ' s'
                elif info < 3600 and info >= 1800:
                    return '{:.2f}'.format(info / 60) + ' min'
                elif info < 86400 and info >= 3600:
                    return '{:.2f}'.format(info / 3600) + ' h'
                elif info >= 86400:
                    return '{:.2f}'.format(info / 86400) + ' d'
                else:
                    logger.error(f'时间解析失败')
            else:
                if info < 1000:
                    return '{:.2f}'.format(info)
                elif info >= 1000 and info < 1000000:
                    return '{:.2f}'.format(info / 1000) + ' k'
                else:
                    return '{:.2f}'.format(info)
        except Exception as e:
            logger.error(f'{e}')
            return str(info)


    def query_total_info(self,ipaddress = ''):

        logger.info(f'开始获取总体汇总情况')
        return_dict = {}
        df = self.get_df_from_ipaddress(ipaddress)
        df_rows = len(df)
        qps, conc = 0, 0

        # 以下是总体汇总情况
        sql_duration_sec = (pd.to_datetime(df['sql_finish_time'], format="%Y/%m/%d", errors='coerce').max() - pd.to_datetime(df['sql_finish_time'], format="%Y/%m/%d", errors='coerce').min()).total_seconds()
        if sql_duration_sec > 0:
            qps = '{:.2f}'.format(df_rows / sql_duration_sec)
            conc = '{:.4f}'.format((df['query_time'].astype(float).sum()) / sql_duration_sec)

        return_dict['qps'] = qps
        return_dict['concurrency'] = conc
        return_dict['total'] = df_rows
        return_dict['unique'] = len(df.groupby('md5_query'))
        return_dict['min_time'] = str(pd.to_datetime(df['sql_finish_time'], format="%Y/%m/%d", errors='coerce').min())
        return_dict['max_time'] = str(pd.to_datetime(df['sql_finish_time'], format="%Y/%m/%d", errors='coerce').max())
        return_dict['total_data'] = {}
        for items in ['query_time', 'lock_time', 'rows_sent', 'rows_examined', 'sql_text']:
            if items == 'sql_text':
                attribute = 'Query size'
            elif items == 'query_time':
                attribute = 'Exec time'
            elif items == 'lock_time':
                attribute = 'Lock time'
            elif items == 'rows_sent':
                attribute = 'Rows sent'
            elif items == 'rows_examined':
                attribute = 'Rows examine'
            else:
                continue
            if items in ['sql_text']:
                data = df[items].str.len()
            else:
                data = df[items].astype(float)

            data_dict = {'sum': self.num_format(data.sum(), items),
                         'min': self.num_format(data.min(), items),
                         'max': self.num_format(data.max(), items),
                         'mean': self.num_format(data.mean(), items),
                         'quantile': self.num_format(data.quantile(), items),
                         'std': self.num_format(data.std(), items),
                         'quantile_95': self.num_format(data.quantile(0.95), items)}

            return_dict['total_data'][attribute] = data_dict

        logger.success(f'获取总体汇总情况完成')
        return return_dict

    def query_profile(self,sort_column = 'query_time',group_column = 'query_id',ipaddress = ''):
        if sort_column == 'query_time':
            sort_column_sql = 'response_time'
        elif sort_column == 'query_count':
            sort_column_sql = 'calls'
        elif sort_column == 'lock_time':
            sort_column_sql = 'lock_time'
        else:
            logger.error(f'query_profile 查询失败 请检查 sort_column: {sort_column} 信息是否正确。')
            sys.exit(1)

        if group_column == 'query_id':
            group_column_sql = 'substr(query, 1, 2000)'
        elif group_column == 'sql_text':
            group_column_sql = 'substr(sql_text, 1, 2000)'
        else:
            logger.error(f'query_profile 查询失败 请检查 group_column: {group_column} 信息是否正确。')
            sys.exit(1)

        new_df = self.get_df_from_ipaddress(ipaddress)

        query = f"""
                    with a as (
                        select sum(query_time) total_query_time,
                               sum(lock_time) total_query_time
                         from new_df
                    )
                    select ROW_NUMBER() over(order by {sort_column_sql} desc) rank,
                           query_id,
                           response_time,
                           response_time / a.total_query_time * 100 percent_info,
                           calls,
                           avg_query_time,
                           lock_time,
                           avg_lock_time,
                           sql_text
                      from (
                            select md5_query query_id, 
                                  {group_column_sql} sql_text,
                                   sum(query_time) response_time,
                                   count(md5_query) calls,
                                   avg(query_time) avg_query_time,
                                   sum(lock_time) lock_time,
                                   avg(lock_time) avg_lock_time
                              from new_df
                             group by {group_column_sql}, md5_query
                            ),a 
                      {self.query_total_size_sql}
                    """
        df_sql = duckdb.sql(query).df()
        return df_sql

    def write_query_profile(self,data,rpt,write_type):
        if write_type == 'txt':
            tplt = "{:>5}\t{:>35}\t{:>15}\t{:>10}\t{:>10}\t{:>10}\t{:>20}\t{:>20}\t{:<20}"
            rpt.write('#' + tplt.format('Rank', 'Query ID', 'Response time', 'Percent', 'Calls', 'R/Call','Lock Time', 'Avg Lock Time', 'SQL_TEXT') + '\n')
            rpt.write('#' + tplt.format('====', '=================', '==============', '==========', '=======', '=========','=========','=============', '===========') + '\n')

            for v_all_data, row in data.iterrows():
                rank = str(row['rank'])
                query_id = str(row['query_id'])
                response_time = self.num_format(row['response_time'])
                percent = str('{:.2f}'.format(row['percent_info']) + '%')
                calls = str(row['calls'])
                avg_query_time = self.num_format(row['avg_query_time'])
                sql_text = str(row['sql_text'])
                sql_text = re.sub(r' +', ' ', sql_text).strip()
                lock_time = self.num_format(row['lock_time'])
                avg_lock_time = self.num_format(row['avg_lock_time'])
                rpt.write('#' + tplt.format(rank, query_id, response_time, percent, calls, avg_query_time, lock_time, avg_lock_time, sql_text) + '\n')

        elif write_type == 'html':
            return_info = []
            for v_all_data, row in data.iterrows():
                rank = str(row['rank'])
                query_id = str(row['query_id'])
                response_time = self.num_format(row['response_time'])
                percent = str('{:.2f}'.format(row['percent_info']) + '%')
                calls = str(row['calls'])
                avg_query_time = self.num_format(row['avg_query_time'])
                sql_text = str(row['sql_text'])
                sql_text = re.sub(r' +', ' ', sql_text).strip()
                lock_time = self.num_format(row['lock_time'])
                avg_lock_time = self.num_format(row['avg_lock_time'])

                return_info.append({
                        'rank': rank,
                        'query_id': query_id,
                        'response_time': response_time,
                        'percent': percent,
                        'calls': calls,
                        'avg_query_time': avg_query_time,
                        'sql_text': sql_text,
                        'lock_time': lock_time,
                        'avg_lock_time': avg_lock_time
                })
            return return_info
        else:
            logger.error(f'write_query_profile 失败 请检查 write_type: {write_type} 信息是否正确。')



    def query_detail_info(self,ipaddress,query_id_list=[]):

        new_df = self.get_df_from_ipaddress(ipaddress = ipaddress)
        if len(query_id_list) > 0:
            query_id_df = pd.DataFrame(query_id_list,columns = ['query_id'])
            query = f"""
                        select max(t1.md5_query) md5_query,max(ipaddress) ipaddress
                          from new_df t1
                          join query_id_df t2
                            on t1.md5_query = t2.query_id
                         group by t1.md5_query 
                         order by sum(query_time) desc
                        """
        else:
            query = f"""
            select max(md5_query) md5_query,max(ipaddress) ipaddress
              from new_df 
             group by md5_query 
             order by sum(query_time) desc
            """
        
        df_sql = duckdb.sql(query).df()
        return df_sql

    def write_query_detail(self, data, rpt, write_type):
        ipaddress = data['ipaddress'].strip()
        sql_md5_str = data['md5_query'].strip()
        df = self.get_df_from_ipaddress(ipaddress)
        df_rows = len(df)

        return_info = {}
        if write_type in ['txt']:
            tplt = "{:>15}\t{:>10}\t{:>10}\t{:>10}\t{:>10}\t{:>10}\t{:>20}\t{:>10}\t{:>10}"
            rpt.write('\n# Query ID: ' + sql_md5_str + '')
            rpt.write('\n# Ipaddress: ' + ipaddress + '')
            rpt.write('\n# User: ' + str(df[df.md5_query.isin([sql_md5_str])]['user'].unique()) + '')
            rpt.write('\n# Client IP: ' + str(df[df.md5_query.isin([sql_md5_str])]['client_ip'].unique()) + '')
            rpt.write('\n# Schema: ' + str(df[df.md5_query.isin([sql_md5_str])]['schema'].unique()) + '')
            rpt.write('\n# Time range: ' + str(df[df.md5_query.isin([sql_md5_str])]['sql_finish_time'].min()) + ' to ' + str(df[df.md5_query.isin([sql_md5_str])]['sql_finish_time'].max()) + '')
            rpt.write('\n#' + tplt.format('Attribute', 'pct', 'total', 'min', 'max', 'avg', '95%', 'stddev', 'median') + '')
            rpt.write('\n#' + tplt.format('============', '===', '=======', '=======', '=======', '=======','=========','=======', '=======') + '')
            rpt.write('\n#' + tplt.format('Count', '{:.0f}'.format(len(df[df.md5_query.isin([sql_md5_str])]) / df_rows), str(len(df[df.md5_query.isin([sql_md5_str])])), '', '', '', '','','') + '\n')


            for col_name in ['query_time', 'lock_time', 'rows_sent', 'rows_examined']:
                if col_name == 'query_time':
                    out_col_name = 'Exec time'
                elif col_name == 'lock_time':
                    out_col_name = 'Lock time'
                elif col_name == 'rows_sent':
                    out_col_name = 'Rows sent'
                elif col_name == 'rows_examined':
                    out_col_name = 'Rows examine'
                else:
                    out_col_name = ''
                info = df[df.md5_query.isin([sql_md5_str])][col_name].astype(float)

                col_sum = self.num_format(info.sum(),col_name)
                col_min = self.num_format(info.min(),col_name)
                col_max = self.num_format(info.max(),col_name)
                col_avg = self.num_format(info.mean(),col_name)
                col_per95 = self.num_format(info.quantile(0.95),col_name)
                col_std = self.num_format(info.std(), col_name)
                col_per = self.num_format(info.quantile(), col_name)

                data1 = info.sum() / df[col_name].astype(float).sum() if df[col_name].astype(float).sum() != 0 else 0
                col_percent = self.num_format(data1,col_name)
                rpt.write('#' + tplt.format( out_col_name,col_percent,col_sum,col_min,col_max,col_avg,col_per95,col_std,col_per)+ '\n')



            if self.if_data:
                rpt.write(df[df.md5_query.isin([sql_md5_str])]['sql_text'].min() + '\n')
            else:
                rpt.write(df[df.md5_query.isin([sql_md5_str])]['query'].min() + '\n')
        elif write_type in ['html']:

            return_info = {
                'query_id': sql_md5_str,
                'ipaddress': ipaddress,
                'user': str(df[df.md5_query.isin([sql_md5_str])]['user'].unique()),
                'client_ip': str(df[df.md5_query.isin([sql_md5_str])]['client_ip'].unique()),
                'time_range': str(df[df.md5_query.isin([sql_md5_str])]['sql_finish_time'].min()) + ' to ' + str(
                    df[df.md5_query.isin([sql_md5_str])]['sql_finish_time'].max()),
                'schema':str(df[df.md5_query.isin([sql_md5_str])]['schema'].unique()),
                'data': {}
            }

            return_info['data']['count'] = {
                'pct': '{:.0f}'.format(len(df[df.md5_query.isin([sql_md5_str])]) / df_rows),
                'total': str(len(df[df.md5_query.isin([sql_md5_str])])),
                'min': '',
                'max': '',
                'avg': '',
                'per95': '',
                'stddev': '',
                'per': ''
            }

            for col_name in ['query_time', 'lock_time', 'rows_sent', 'rows_examined']:
                info = df[df.md5_query.isin([sql_md5_str])][col_name].astype(float)

                col_sum = self.num_format(info.sum(), col_name)
                col_min = self.num_format(info.min(), col_name)
                col_max = self.num_format(info.max(), col_name)
                col_avg = self.num_format(info.mean(), col_name)
                col_per95 = self.num_format(info.quantile(0.95), col_name)
                col_std = self.num_format(info.std(), col_name)
                col_per = self.num_format(info.quantile(), col_name)

                data1 = info.sum() / df[col_name].astype(float).sum() if df[col_name].astype(float).sum() != 0 else 0
                col_percent = self.num_format(data1, col_name)

                return_info['data'][col_name] = {
                    'pct': col_percent,
                    'total': col_sum,
                    'min': col_min,
                    'max': col_max,
                    'avg': col_avg,
                    'per95': col_per95,
                    'stddev': col_std,
                    'per': col_per
                }

            if self.if_data:
                return_info['data']['sql_text'] = df[df.md5_query.isin([sql_md5_str])]['sql_text'].min()
            else:
                return_info['data']['sql_text'] = df[df.md5_query.isin([sql_md5_str])]['query'].min()

            return return_info
        else:
            logger.error(f'write_query_detail 失败 请检查 write_type: {write_type} 信息是否正确。')
    def txt_query_detail_report(self, rpt, ipaddress = ''):
        logger.info(f'开始生成{ipaddress}慢日志详情')

        # 以下是每个SQL的详细情况
        detail_info  = self.query_detail_info(ipaddress)
        rpt.write('\n' * 3)
        rpt.write(f'#' * 100 + '\n')
        if ipaddress != '':
            title = f'# {ipaddress} Parse Detail Report'
        else:
            title = f'# Total Parse Detail Report'
        rpt.write(f'{title}' + ' ' * (100 - len(title) - 1) + '#\n')
        rpt.write(f'#' * 100 + '')
        for i, row in detail_info.iterrows():
            self.write_query_detail(row, rpt, 'txt')

    def txt_query_report(self, rpt, ipaddress = ''):

        logger.info(f'开始生成{ipaddress}慢日志汇总报告')

        rpt.write('\n' * 3)
        rpt.write(f'#' * 100 + '\n')
        if ipaddress != '':
            title = f'# {ipaddress} Parse Report'
        else:
            title = f'# Total Parse Report'
        rpt.write(f'{title}' + ' ' * (100 - len(title) - 1) + '#\n')
        rpt.write(f'#' * 100 + '\n')

        rpt.write(f'# Current date: {str(datetime.datetime.now())}\n')
        rpt.write(f'# Hostname: {str(platform.node())}\n')
        rpt.write(f"# Ipaddress: {str(get_ip_address()) if ipaddress == '' else ipaddress}\n")

        if self.parse_type == 1:
            rpt.write(f'# Files Directory: {self.slow_log_path}\n')
            rpt.write(f'# Files Name: {self.slow_log_name}\n')

        if self.parse_type == 2:
            rpt.write(f'# Data From: Get Data From ElasticSearch\n')

        total_info_dict = self.query_total_info(ipaddress)
        rpt.write(f"# Overall: {total_info_dict['total']} total, {str(total_info_dict['unique'])} unique, {str(total_info_dict['qps'])} QPS, {str(total_info_dict['concurrency'])} concurrency\n")
        rpt.write(f"# Time range: {total_info_dict['min_time']} to {total_info_dict['max_time']}\n")

        # 设置格式tplt，20代表间隔距离，可根据自己需要调整
        tplt = "{:>20}\t{:>15}\t{:>15}\t{:>15}\t{:>15}\t{:>15}\t{:>15}\t{:>15}"
        rpt.write('#' + tplt.format('Attribute', 'total', 'min', 'max', 'avg', '95%', 'stddev', 'median') + '\n')
        rpt.write('#' + tplt.format('=================', '===========', '===========', '===========', '===========', '===========', '===========', '===========') + '\n')
        for key,value in total_info_dict['total_data'].items():
            rpt.write('#' + tplt.format(key, value['sum'], value['min'], value['max'], value['mean'], value['quantile_95'], value['std'], value['quantile']) + '\n')

        rpt.write('\n# Query Time Sort Profile(By QUERY_ID)\n')
        data = self.query_profile('query_time', 'query_id',ipaddress)
        self.write_query_profile(data,rpt,'txt')

        if self.if_data:
            rpt.write('\n# Query Time Sort Profile(By SQL_TEXT)\n')
            data = self.query_profile('query_time', 'sql_text',ipaddress)
            self.write_query_profile(data, rpt, 'txt')

        rpt.write('\n# Calls Sort Profile(By QUERY_ID)\n')
        data = self.query_profile('query_count', 'query_id',ipaddress)
        self.write_query_profile(data, rpt, 'txt')

        if self.if_data:
            rpt.write('\n# Calls Sort Profile(By SQL_TEXT)\n')
            data = self.query_profile('query_count', 'sql_text',ipaddress)
            self.write_query_profile(data, rpt, 'txt')

        rpt.write('\n# Lock Time Sort Profile(By QUERY_ID)\n')
        data = self.query_profile('lock_time', 'query_id',ipaddress)
        self.write_query_profile(data, rpt, 'txt')

        if self.if_data:
            rpt.write('\n# Lock Time Sort Profile(By SQL_TEXT)\n')
            data = self.query_profile('lock_time', 'sql_text',ipaddress)
            self.write_query_profile(data, rpt, 'txt')

        logger.success(f'生成{ipaddress}慢日志汇总报告')
    def query_max_exec_time(self):
        new_df = self.df
        # 分组汇总输出sql执行情况按最大执行耗时排序
        query = """
        select row_number() over(order by max_tim desc) 序号,
               cast(s as varchar) SQL文本,
               max_tim 最大执行耗时_秒,
               min_tim 最小执行耗时_秒,
               avg_90 百分之90平均执行耗时_秒,
               avg_tim 平均执行耗时_秒,
               cnt 执行次数,
               max_rcount 行数,
               lock_time 等待时间_秒
          from (select s,
                       max(tim) over(partition by s) max_tim,
                       min(tim) over(partition by s) min_tim,
                       avg(tim) over(partition by s) avg_tim,
                       avg(case when tile <= 9 then tim else null end) over(partition by s) avg_90,
                       count(1) over(partition by s) cnt,
                       row_number() over(partition by s order by tim desc) r,
                       max(rows_examined) over(partition by s) max_rcount,
                       lock_time
                  from (select sql_text s,
                               cast(query_time as numeric) tim,
                               ntile(10) over(partition by sql_text order by query_time) tile,
                               cast(rows_examined as numeric) rows_examined,
                               cast(lock_time as numeric) lock_time
                          from new_df
                         where query_time >= 0) t1) xx
         where r = 1
        """
        df_sql = duckdb.sql(query).df()
        fields = ['序号', 'sql文本', '最大执行耗时_秒', '最小执行耗时_秒', '百分之90平均执行耗时_秒', '平均执行耗时_秒', '执行次数', '行数', '等待时间_秒']  # 获取所有字段名
        all_data = df_sql.apply(lambda x: tuple(x), axis=1).values.tolist()  # 每一行转为元组tuple，所有数据以列表list输出
        excel_rows_num = len(all_data) + 1
        return all_data, fields, excel_rows_num

    def query_max_exec_count(self):
        new_df = self.df
        # 分组汇总输出sql执行情况按最大执行次数排序
        query = """
        select row_number() over(order by cnt desc) 序号,
               cast(s as varchar) SQL文本,
               max_tim 最大执行耗时_秒,
               min_tim 最小执行耗时_秒,
               avg_90 百分之90平均执行耗时_秒,
               avg_tim 平均执行耗时_秒,
               cnt 执行次数,
               max_rcount 行数,
               lock_time 等待时间_秒
          from (select s,
                       max(tim) over(partition by s) max_tim,
                       min(tim) over(partition by s) min_tim,
                       avg(tim) over(partition by s) avg_tim,
                       avg(case when tile <= 9 then tim else null end) over(partition by s) avg_90,
                       count(1) over(partition by s) cnt,
                       row_number() over(partition by s order by tim desc) r,
                       max(rows_examined) over(partition by s) max_rcount,
                       lock_time
                  from (select sql_text s,
                               cast(query_time as numeric) tim,
                               ntile(10) over(partition by sql_text order by query_time) tile,
                               cast(rows_examined as numeric) rows_examined,
                               cast(lock_time as numeric) lock_time
                          from new_df
                         where query_time >= 0) t1) xx
         where r = 1
           and cnt >= 0
        """
        df_sql = duckdb.sql(query).df()
        fields = ['序号', 'sql文本', '最大执行耗时_秒', '最小执行耗时_秒', '百分之90平均执行耗时_秒', '平均执行耗时_秒', '执行次数', '行数', '等待时间_秒']  # 获取所有字段名
        all_data = df_sql.apply(lambda x: tuple(x), axis=1).values.tolist()  # 所有数据
        excel_rows_num = len(all_data) + 1
        # print(excel_rows_num)
        return all_data, fields, excel_rows_num

    def query_all_data(self):
        new_df = self.df
        query = """
        select row_number() over() as row_num,
               sql_finish_time,
               query_time * 1000 as query_time,
               exec_client,
               substr(replace(replace(cast(sql_text as varchar),chr(10),''),chr(13),''),1,200) as sub_txt,
               optype
          from new_df
        """
        df_sql = duckdb.sql(query).df()
        # all_data = df_sql.apply(lambda x: tuple(x), axis=1).values.tolist()
        df_data_list = df_sql.values.tolist()
        return df_data_list


def check_report_dir():
    report_dir_name = f'slowlog_report_{time_day_format}'

    output_result_dir = ReadConfig.get_path('output_dir')
    print(output_result_dir)
    if len(output_result_dir) <= 0 or output_result_dir == '.':
        output_result_dir = os.path.join(exepath,'slowlog_report')

    log_dir = ReadConfig.get_path('log_dir')
    if len(log_dir) <= 0 or log_dir == '.':
        log_dir = os.path.join(exepath,'log')

    if not os.path.exists(log_dir):
        os.makedirs(log_dir)

    report_dir = os.path.join(output_result_dir, report_dir_name)
    set_logger(log_dir)

    if not os.path.exists(report_dir):
        os.makedirs(report_dir)
        logger.success(f"创建输出结果目录 {report_dir} 成功！")
    else:
        logger.info(f"输出结果目录 {report_dir} 已存在")

    return report_dir




def run():
    print_info()
    report_dir = check_report_dir()

    slow_log_info = slowlog()
    out_df = slow_log_info.run()


    query_report_info = query_report(out_df, report_dir)
    query_report_info.run()
    time_end = time.time()

    logger.success(f"程序运行共耗时：{'{:.2f}'.format(time_end - time_start_1)} 秒")


if __name__ == "__main__":
    run()

